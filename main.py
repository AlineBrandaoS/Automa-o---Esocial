import time                                                                   
import os                                       
import undetected_chromedriver as uc          
from selenium.webdriver.common.by import By       
from selenium.webdriver.chrome.options import Options     
from selenium.webdriver.support.ui import WebDriverWait             
from selenium.webdriver.support import expected_conditions as EC          
from openpyxl import load_workbook                                               
import glob                                                                       

# --- CONFIGURAÇÃO DA PASTA DE DOWNLOAD ---
default_download_dir = r"C:\Users\Pierre\Downloads"  # Essa variavel é utilizada para dizer onde o arquivo irá ser baixado.

# --- CONFIGURAÇÃO DAS OPÇÕES DO CHROME ---
chrome_options = Options()                     
prefs = {                                                 
    "download.default_directory": default_download_dir,   
    "download.prompt_for_download": False,                    
    "download.directory_upgrade": True,                       
    "safebrowsing.enabled": True                             
}
chrome_options.add_experimental_option("prefs", prefs)

# --- FUNÇÕES AUXILIARES ---

def esperar_download_e_retornar_caminho(arquivos_iniciais, timeout=120, download_folder=default_download_dir):
    # Essa função foi utilizada para que ele espere que o download seja concluido completamente para não ter erro
    # com o arquivo
    tempo_inicial = time.time()
    
    while time.time() - tempo_inicial < timeout:
        arquivos_atuais = set(glob.glob(os.path.join(download_folder, '*')))
        novos_arquivos = arquivos_atuais - arquivos_iniciais
        
        novos_arquivos_nao_temp = [f for f in novos_arquivos if not f.endswith(('.tmp', '.crdownload'))]

        if len(novos_arquivos_nao_temp) > 0:
            caminho_final_arquivo = novos_arquivos_nao_temp[0]
            
            tamanho_anterior = -1
            while True:
                time.sleep(1)
                tamanho_atual = os.path.getsize(caminho_final_arquivo)
                if tamanho_atual == tamanho_anterior:
                    return caminho_final_arquivo
                tamanho_anterior = tamanho_atual
        
        time.sleep(1)
    return None

def identificar_tipo_de_registro(driver):
    # Essa foi criada para que ele identifique o elemento que esta sendo utilizado
    # FGTS/INSS para que não ocorra de fazer o download do XML errado.
    
    xpath_fgts = '//*[@id="Cpf"]'
    xpath_inss = '//*[@id="CpfPesquisa"]'

    try:
        if driver.find_elements(By.XPATH, xpath_fgts):
            return "FGTS"
        elif driver.find_elements(By.XPATH, xpath_inss):
            return "INSS"
        else:
            return "Outro"
            
    except Exception:
        return "Não Identificado"


# --- FUNÇÃO PRINCIPAL ---
def task():
    print("\n--- INÍCIO DA TAREFA DE AUTOMAÇÃO ---")
    
    # --- DEFINA O TIPO DE PESQUISA AQUI ---
    TIPO_DE_PESQUISA = "INSS"

    if TIPO_DE_PESQUISA.upper() == "FGTS":
        xpath_campo_cpf = '//*[@id="Cpf"]'
        xpath_botao_pesquisar = '//*[@id="conteudo-pagina"]/form/div/section/div/div[3]/input'
        xpath_campo_periodo = '//*[@id="PeriodoApuracao"]'
    elif TIPO_DE_PESQUISA.upper() == "INSS":
        xpath_campo_cpf = '//*[@id="CpfPesquisa"]'
        xpath_botao_pesquisar = '//*[@id="form-totalizador"]/div/section/input'
        xpath_campo_periodo = '//*[@id="PeriodoApuracaoPesquisa"]'
    else:
        print("❌ Erro: TIPO_DE_PESQUISA deve ser 'FGTS' ou 'INSS'.")
        return

    driver = None
    cpfs_com_erro = []
    
    # --- PLANILHA ---
    # Nessa parte do codigo ele lê todos os dados da planilha e identifica ela 
    #### OBS: A SUA PLANILHA TEM QUE ESTÁ JUNTO COM A PASTA DE CODIGO!!
    # Nesse caso é uma planilha de exemplo.
    try:
        print("📖 Carregando a planilha do Excel...")
        planilha_efetivo = load_workbook('EFETIVO - EXEMPLO.xlsx')
        pagina_efetivo = planilha_efetivo['EFETIVO']
        print("✅ Planilha carregada com sucesso.")
    except FileNotFoundError:
        print("❌ Erro: Arquivo 'EFETIVO - EXEMPLO.xlsx' não encontrado.")
        return
    
    # --- LÓGICA DE INÍCIO DA SESSÃO ---
    ## Bom, aqui eu obtive meu principal problema, em entrar no esocial sem o site identificar que sou um bot :3
    ## então, resolvi fazendo manualmente.
    try:
        print("🚀 Inicializando o navegador...")
        driver = uc.Chrome(options=chrome_options, use_subprocess=True, read_timeout=180)
        driver.maximize_window()
        driver.get("https://login.esocial.gov.br/login.aspx")
        print("✅ Navegador iniciado. Por favor, faça o login e navegue para a tela de consulta.")
        input("⚠️ Pressione ENTER para continuar após o login...")
    ## inserir o periodo.
        campo_periodo = WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, xpath_campo_periodo)))
        campo_periodo.clear()
        campo_periodo.send_keys("07/2025")
        time.sleep(1)
    except Exception as e:
        print(f"❌ Erro fatal ao iniciar o navegador: {e}")
        if driver: driver.quit()
        return

    # Loop pelos CPFs
    for linha in range(2, pagina_efetivo.max_row + 1):
        cpf = pagina_efetivo.cell(row=linha, column=2).value
        
        if cpf is None:
            continue

        print(f"\n--- Processando CPF da linha {linha}: {cpf} ---")
        try:
            campo_cpf = WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, xpath_campo_cpf)))
            campo_cpf.clear()
            campo_cpf.send_keys(str(cpf))
            
            print("⏳ Dando uma pausa para a página carregar a validação...")
            time.sleep(5)
            
            botao_pesquisar = WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, xpath_botao_pesquisar)))
            botao_pesquisar.click()

            # A identificação do tipo de registro agora é feita depois da pesquisa, para saber o resultado
            tipo_de_registro_encontrado = identificar_tipo_de_registro(driver)
            print(f"✅ O script identificou um registro de: {tipo_de_registro_encontrado}")
            
            try:
                arquivos_iniciais = set(glob.glob(os.path.join(default_download_dir, '*')))
                
                botao_baixar = WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.PARTIAL_LINK_TEXT, "Baixar XML")))
                botao_baixar.click()
                print("⏳ Aguardando a conclusão do download...")
                
                caminho_arquivo_baixado = esperar_download_e_retornar_caminho(arquivos_iniciais)
                
                if caminho_arquivo_baixado:
                    print(f"✅ Download finalizado para CPF {cpf}.")
                    print(f"✅ Arquivo salvo em: {caminho_arquivo_baixado}")
                else:
                    print(f"⚠️ Aviso: Timeout ao aguardar download para CPF {cpf}.")
                    cpfs_com_erro.append(cpf)
                    
            except Exception as e:
                print(f"❗ O CPF {cpf} pode não ter tempo de empresa para o FGTS ser depositado e o XML não aparecer. Detalhes: {e}")
                cpfs_com_erro.append(cpf)
                
        except Exception as e:
            print(f"❌ Ocorreu um erro inesperado ao processar o CPF {cpf}. Detalhes: {e}")
            cpfs_com_erro.append(cpf)
            
    print("\n--- FIM DA TAREFA DE AUTOMAÇÃO ---")
    print(f"\n🎉 Todos os downloads concluídos (ou tentados)!")
    # aqui ele mostra os cpfs que deram errado em uma lista.
    if len(cpfs_com_erro) > 0:
        print(f"\n❌ Atenção: {len(cpfs_com_erro)} CPFs apresentaram erro.")
        print("CPFs com erro:")
        for cpf_erro in cpfs_com_erro:
            print(f" - {cpf_erro}")
    else:
        print("\n✅ Todos os CPFs foram processados com sucesso, sem erros.")
        
    if driver:
        driver.quit()

    task()
